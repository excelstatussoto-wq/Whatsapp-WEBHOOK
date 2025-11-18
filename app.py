from flask import Flask, request, Response, abort
import os, datetime, re
import threading

USE_SHEETS = os.getenv("USE_SHEETS", "false").lower() in ("1","true","yes")
LOCK = threading.Lock()

app = Flask(__name__)

def extract_code(text):
    m = re.search(r"\b[A-Z0-9\-]{3,}\b", text, re.I)
    return m.group(0) if m else text.strip()

if USE_SHEETS:
    # Google Sheets setup (requires service account JSON file path in CREDS_FILE and SHEET_KEY env var)
    import gspread
    CREDS_FILE = os.getenv("CREDS_FILE", "service_account.json")
    SHEET_KEY = os.getenv("SHEET_KEY", "")
    if not SHEET_KEY:
        raise RuntimeError("SHEET_KEY environment variable required when USE_SHEETS=true")
    gc = gspread.service_account(filename=CREDS_FILE)
    sh = gc.open_by_key(SHEET_KEY)
    worksheet = sh.sheet1
else:
    # Excel local setup
    from openpyxl import Workbook, load_workbook
    EXCEL_PATH = os.getenv("EXCEL_PATH", "whatsapp_codes.xlsx")
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp","From","Code","RawMessage"])
        wb.save(EXCEL_PATH)

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.form.get("Body") or (request.json and request.json.get("Body"))
    sender = request.form.get("From") or (request.json and request.json.get("From"))
    if not body:
        abort(400, "No message body")
    code = extract_code(body)
    timestamp = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"

    if USE_SHEETS:
        # Append to Google Sheet
        worksheet.append_row([timestamp, sender, code, body], value_input_option="USER_ENTERED")
    else:
        # Write to local Excel with lock
        with LOCK:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            ws.append([timestamp, sender, code, body])
            wb.save(EXCEL_PATH)

    return Response(status=200)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)))
